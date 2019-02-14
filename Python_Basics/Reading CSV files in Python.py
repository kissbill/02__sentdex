from openpyxl import Workbook
from openpyxl import load_workbook

wb = Workbook()


wb = load_workbook('noname.xlsx')
wb = wb.active
sheet_ranges = wb['Sheet1']

print(sheet_ranges['J10'].value)
print(sheet_ranges['D10'].value)
print(sheet_ranges['L9'].value)

=HA('C:\O1\O1.02.eCar\eCarIntern\temp\evocszs2\[szabi alap.xlsx]Sheet1'!L11;0;1)